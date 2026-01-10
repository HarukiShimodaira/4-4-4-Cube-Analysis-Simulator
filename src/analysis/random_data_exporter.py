"""
Random Data Excel Exporter
Exports collected random operation data to Excel format

Copyright (c) 2025 Haruki Shimodaira
Licensed under the MIT License - see LICENSE file for details
"""

from typing import Dict, Any, List, Optional
import math
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import io


class RandomDataExcelExporter:
    """Export random operation data to Excel format"""
    
    @staticmethod
    def export_to_excel(data: Dict[str, Any], filename: Optional[str] = None) -> Optional[bytes]:
        """
        Export collected data to Excel file
        
        Args:
            data: Data dictionary from RandomDataCollector.collect_data()
            filename: Optional filename to save (if None, returns bytes)
            
        Returns:
            bytes: Excel file content if filename is None
        """
        wb = openpyxl.Workbook()
        
        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        # Create sheets
        RandomDataExcelExporter._create_metadata_sheet(wb, data)
        RandomDataExcelExporter._create_initial_state_sheet(wb, data)
        RandomDataExcelExporter._create_position_sheet(wb, data)
        RandomDataExcelExporter._create_rotation_matrix_sheet(wb, data)
        RandomDataExcelExporter._create_rotation_vector_sheet(wb, data)
        
        # Save or return bytes
        if filename:
            wb.save(filename)
            return None
        else:
            # Return as bytes
            virtual_file = io.BytesIO()
            wb.save(virtual_file)
            virtual_file.seek(0)
            return virtual_file.read()
    
    @staticmethod
    def _create_metadata_sheet(wb: openpyxl.Workbook, data: Dict[str, Any]):
        """Create metadata sheet with experiment info"""
        ws = wb.create_sheet("メタデータ", 0)
        
        # Header style
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        # Metadata
        ws['A1'] = "項目"
        ws['B1'] = "値"
        ws['A1'].fill = header_fill
        ws['B1'].fill = header_fill
        ws['A1'].font = header_font
        ws['B1'].font = header_font
        
        metadata = [
            ("生成日時", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
            ("乱数シード", data['seed']),
            ("操作回数", len(data['operation_sequence'])),
            ("キューブレット数", 56),
            ("操作種類", "R, Rp, L, Lp, U, Up, D, Dp, F, Fp, B, Bp, r, rp, l, lp, u, up, d, dp, f, fp, b, bp"),
        ]
        
        for i, (key, value) in enumerate(metadata, start=2):
            ws[f'A{i}'] = key
            ws[f'B{i}'] = value
        
        # Column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 80
    
    @staticmethod
    def _create_initial_state_sheet(wb: openpyxl.Workbook, data: Dict[str, Any]):
        """Create sheet with static initial state data"""
        ws = wb.create_sheet("初期状態")
        
        # Header style
        header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Headers
        headers = ["ID", "タイプ", "初期X", "初期Y", "初期Z"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(1, col, header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        
        # Data
        for cubelet_data in data['initial_state']:
            row = cubelet_data['id'] + 2
            ws.cell(row, 1, cubelet_data['id']).border = border
            ws.cell(row, 2, cubelet_data['type']).border = border
            ws.cell(row, 3, cubelet_data['initial_pos_x']).border = border
            ws.cell(row, 4, cubelet_data['initial_pos_y']).border = border
            ws.cell(row, 5, cubelet_data['initial_pos_z']).border = border
        
        # Column widths
        for col in range(1, 6):
            ws.column_dimensions[get_column_letter(col)].width = 12
    
    @staticmethod
    def _create_position_sheet(wb: openpyxl.Workbook, data: Dict[str, Any]):
        """Create sheet with position data (1 row per operation)"""
        ws = wb.create_sheet("座標データ")
        
        # Styles
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Header row
        ws.cell(1, 1, "操作").fill = header_fill
        ws.cell(1, 1).font = header_font
        ws.cell(1, 1).border = border
        
        # Create headers for all 56 cubelets (X, Y, Z for each)
        col = 2
        for cubelet_id in range(56):
            for axis in ['X', 'Y', 'Z']:
                cell = ws.cell(1, col, f"ID{cubelet_id}_{axis}")
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border
                cell.alignment = Alignment(horizontal='center')
                col += 1
        
        # Data rows (one per snapshot)
        for snapshot_idx, snapshot in enumerate(data['snapshots']):
            row = snapshot_idx + 2
            
            # Operation label
            operation_label = "初期状態" if snapshot.step == 0 else snapshot.operation
            ws.cell(row, 1, operation_label).border = border
            
            # Position data for all cubelets
            col = 2
            for cubelet_data in snapshot.cubelets_data:
                ws.cell(row, col, cubelet_data['pos_x']).border = border
                ws.cell(row, col).number_format = '0.0000'
                col += 1
                ws.cell(row, col, cubelet_data['pos_y']).border = border
                ws.cell(row, col).number_format = '0.0000'
                col += 1
                ws.cell(row, col, cubelet_data['pos_z']).border = border
                ws.cell(row, col).number_format = '0.0000'
                col += 1
        
        # Column widths
        ws.column_dimensions['A'].width = 15
        for col in range(2, 2 + 56 * 3):
            ws.column_dimensions[get_column_letter(col)].width = 10
        
        # Freeze panes
        ws.freeze_panes = 'B2'
    
    @staticmethod
    def _create_rotation_matrix_sheet(wb: openpyxl.Workbook, data: Dict[str, Any]):
        """Create sheet with rotation matrix data (1 row per operation)"""
        ws = wb.create_sheet("回転行列")
        
        # Styles
        header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Header row
        ws.cell(1, 1, "操作").fill = header_fill
        ws.cell(1, 1).font = header_font
        ws.cell(1, 1).border = border
        
        # Create headers for all 56 cubelets (9 matrix elements for each)
        col = 2
        for cubelet_id in range(56):
            for i in range(3):
                for j in range(3):
                    cell = ws.cell(1, col, f"ID{cubelet_id}_M{i}{j}")
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center')
                    col += 1
        
        # Data rows (one per snapshot)
        for snapshot_idx, snapshot in enumerate(data['snapshots']):
            row = snapshot_idx + 2
            
            # Operation label
            operation_label = "初期状態" if snapshot.step == 0 else snapshot.operation
            ws.cell(row, 1, operation_label).border = border
            
            # Rotation matrix data for all cubelets
            col = 2
            for cubelet_data in snapshot.cubelets_data:
                for i in range(3):
                    for j in range(3):
                        key = f'rot_m{i}{j}'
                        ws.cell(row, col, cubelet_data[key]).border = border
                        ws.cell(row, col).number_format = '0.0000'
                        col += 1
        
        # Column widths
        ws.column_dimensions['A'].width = 15
        for col in range(2, 2 + 56 * 9):
            ws.column_dimensions[get_column_letter(col)].width = 10
        
        # Freeze panes
        ws.freeze_panes = 'B2'
    
    @staticmethod
    def _create_rotation_vector_sheet(wb: openpyxl.Workbook, data: Dict[str, Any]):
        """Create sheet with rotation vector data (axis-angle, 1 row per operation)"""
        ws = wb.create_sheet("向きベクトル")
        
        # Styles
        header_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        header_font = Font(bold=True, color="000000")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Header row
        ws.cell(1, 1, "操作").fill = header_fill
        ws.cell(1, 1).font = header_font
        ws.cell(1, 1).border = border
        
        # Create headers for all 56 cubelets (angle[deg] + axis_x, axis_y, axis_z + polar coords for each)
        col = 2
        for cubelet_id in range(56):
            for component in ['角度(°)', '軸X', '軸Y', '軸Z', 'r', 'θ(°)', 'φ(°)']:
                cell = ws.cell(1, col, f"ID{cubelet_id}_{component}")
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border
                cell.alignment = Alignment(horizontal='center')
                col += 1
        
        # Data rows (one per snapshot)
        import numpy as np
        for snapshot_idx, snapshot in enumerate(data['snapshots']):
            row = snapshot_idx + 2
            
            # Operation label
            operation_label = "初期状態" if snapshot.step == 0 else snapshot.operation
            ws.cell(row, 1, operation_label).border = border
            
            # Rotation vector data for all cubelets
            col = 2
            for cubelet_data in snapshot.cubelets_data:
                # Use angle in degrees for readability
                ws.cell(row, col, cubelet_data['rotation_angle']).border = border
                ws.cell(row, col).number_format = '0.00'
                col += 1
                ws.cell(row, col, cubelet_data['rotation_axis_x']).border = border
                ws.cell(row, col).number_format = '0.000000'
                col += 1
                ws.cell(row, col, cubelet_data['rotation_axis_y']).border = border
                ws.cell(row, col).number_format = '0.000000'
                col += 1
                ws.cell(row, col, cubelet_data['rotation_axis_z']).border = border
                ws.cell(row, col).number_format = '0.000000'
                col += 1
                # Polar coordinates of current position
                x = cubelet_data.get('pos_x', 0.0)
                y = cubelet_data.get('pos_y', 0.0)
                z = cubelet_data.get('pos_z', 0.0)
                r = math.sqrt(x * x + y * y + z * z)
                theta = math.degrees(math.atan2(y, x))  # azimuth
                phi = math.degrees(math.acos(z / r)) if r > 1e-9 else 0.0  # inclination from +Z
                ws.cell(row, col, r).border = border
                ws.cell(row, col).number_format = '0.000'
                col += 1
                ws.cell(row, col, theta).border = border
                ws.cell(row, col).number_format = '0.00'
                col += 1
                ws.cell(row, col, phi).border = border
                ws.cell(row, col).number_format = '0.00'
                col += 1
        
        # Column widths
        ws.column_dimensions['A'].width = 15
        for col in range(2, 2 + 56 * 7):
            ws.column_dimensions[get_column_letter(col)].width = 12
        
        # Freeze panes
        ws.freeze_panes = 'B2'
    

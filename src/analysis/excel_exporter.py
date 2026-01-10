"""
Excel Exporter Module
Exports analysis results to Excel format with multiple sheets

Copyright (c) 2025 Haruki Shimodaira
Licensed under the MIT License - see LICENSE file for details
"""

from typing import List, Dict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
import pandas as pd
from .position_analyzer import PairAnalysis, PositionStats
from .edge_tracker import EdgeInfo


class ExcelExporter:
    """
    Exports cube analysis results to Excel format
    """
    
    def __init__(self):
        """Initialize the Excel exporter"""
        self.wb = None
        
    def create_workbook(self) -> Workbook:
        """Create a new Excel workbook"""
        self.wb = Workbook()
        # Remove default sheet
        if 'Sheet' in self.wb.sheetnames:
            self.wb.remove(self.wb['Sheet'])
        return self.wb
    
    def add_summary_sheet(self, trial_data: List[Dict], operations_range: tuple):
        """
        Add summary statistics sheet
        
        Args:
            trial_data: List of trial results
            operations_range: (min_ops, max_ops) tuple
        """
        ws = self.wb.create_sheet("Summary", 0)
        
        # Title
        ws['A1'] = "4×4×4 Cube Analysis Simulator - エッジペア分析サマリー"
        ws['A1'].font = Font(size=14, bold=True)
        ws.merge_cells('A1:D1')
        
        # Basic info
        row = 3
        info = [
            ("総トライアル数:", len(trial_data)),
            ("操作数範囲:", f"{operations_range[0]}-{operations_range[1]}"),
            ("", ""),
        ]
        
        for label, value in info:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
        
        # Calculate aggregated statistics
        if trial_data:
            avg_separated = sum(t['separated_pairs'] for t in trial_data) / len(trial_data)
            avg_distance = sum(t['avg_pair_distance'] for t in trial_data) / len(trial_data)
            avg_same_face = sum(t['pairs_on_same_face'] for t in trial_data) / len(trial_data)
            
            row += 1
            stats = [
                ("平均統計:", ""),
                ("  分離ペア数:", f"{avg_separated:.2f} / 12"),
                ("  平均ペア間距離:", f"{avg_distance:.2f}"),
                ("  同一面ペア数:", f"{avg_same_face:.2f}"),
            ]
            
            for label, value in stats:
                ws[f'A{row}'] = label
                ws[f'B{row}'] = value
                ws[f'A{row}'].font = Font(bold=True)
                row += 1
        
        # Column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
    
    def add_pair_distances_sheet(self, trial_data: List[Dict]):
        """
        Add sheet with pair distances for each trial
        
        Args:
            trial_data: List of trial results with pair_analyses
        """
        ws = self.wb.create_sheet("Pair Distances")
        
        # Headers
        headers = ['トライアル', '操作数', 'UF', 'UR', 'UB', 'UL', 'DF', 'DR', 'DB', 'DL', 
                   'FR', 'FL', 'BR', 'BL', '平均距離', '分離数']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
        
        # Data rows
        for trial_idx, trial in enumerate(trial_data, 1):
            row = trial_idx + 1
            ws.cell(row=row, column=1, value=trial_idx)
            ws.cell(row=row, column=2, value=trial['num_operations'])
            
            # Pair distances
            pair_analyses = trial['pair_analyses']
            pair_dict = {p.pair_id: p.current_distance for p in pair_analyses}
            
            for col, pair_id in enumerate(['UF', 'UR', 'UB', 'UL', 'DF', 'DR', 'DB', 'DL', 
                                           'FR', 'FL', 'BR', 'BL'], 3):
                distance = pair_dict.get(pair_id, 0)
                cell = ws.cell(row=row, column=col, value=round(distance, 2))
                
                # Color code by distance (統一基準: near≤1.5, medium≤3.5, far>3.5)
                if distance > 3.5:
                    # far: 赤
                    cell.fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
                elif distance > 1.5:
                    # medium: 黄色
                    cell.fill = PatternFill(start_color="FFD93D", end_color="FFD93D", fill_type="solid")
                # near (≤1.5): 色なし（デフォルト）
            
            ws.cell(row=row, column=15, value=round(trial['avg_pair_distance'], 2))
            ws.cell(row=row, column=16, value=trial['separated_pairs'])
        
        # Auto-width columns
        for col in range(1, 17):
            ws.column_dimensions[chr(64 + col)].width = 10
    
    def add_raw_data_sheet(self, trial_data: List[Dict]):
        """
        Add raw position data for all trials and all edges
        Each row is one edge in one trial
        
        Args:
            trial_data: List of trial results
        """
        ws = self.wb.create_sheet("Raw Data")
        
        # Headers
        headers = ['トライアル', '操作数', 'エッジID', 'ペアID', 
                   '初期X', '初期Y', '初期Z', 
                   '現在X', '現在Y', '現在Z', 
                   '移動距離', 'ペア間距離', '距離区分', '同一面']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
        
        # Data rows - all edges for all trials
        current_row = 2
        for trial_idx, trial in enumerate(trial_data, 1):
            edges = trial['edges']
            pair_analyses = trial['pair_analyses']
            
            # Validate edge count
            if len(edges) != 24:
                raise ValueError(f"Trial {trial_idx} has {len(edges)} edges, expected 24")
            if len(pair_analyses) != 12:
                raise ValueError(f"Trial {trial_idx} has {len(pair_analyses)} pairs, expected 12")
            
            # Create pair distance lookup
            pair_dist_dict = {p.pair_id: (p.current_distance, p.distance_category, p.same_face) 
                             for p in pair_analyses}
            
            for edge in edges:
                ws.cell(row=current_row, column=1, value=trial_idx)
                ws.cell(row=current_row, column=2, value=trial['num_operations'])
                ws.cell(row=current_row, column=3, value=edge.edge_id)
                ws.cell(row=current_row, column=4, value=edge.pair_id)
                
                # Initial position
                ws.cell(row=current_row, column=5, value=round(edge.initial_position.x, 2))
                ws.cell(row=current_row, column=6, value=round(edge.initial_position.y, 2))
                ws.cell(row=current_row, column=7, value=round(edge.initial_position.z, 2))
                
                # Current position
                ws.cell(row=current_row, column=8, value=round(edge.current_position.x, 2))
                ws.cell(row=current_row, column=9, value=round(edge.current_position.y, 2))
                ws.cell(row=current_row, column=10, value=round(edge.current_position.z, 2))
                
                # Movement distance
                ws.cell(row=current_row, column=11, value=round(edge.moved_distance, 2))
                
                # Pair distance and categorization - must exist
                if edge.pair_id not in pair_dist_dict:
                    raise ValueError(f"Trial {trial_idx}: Pair {edge.pair_id} not found in analyses")
                
                pair_dist, dist_cat, same_face = pair_dist_dict[edge.pair_id]
                ws.cell(row=current_row, column=12, value=round(pair_dist, 2))
                ws.cell(row=current_row, column=13, value=dist_cat)
                ws.cell(row=current_row, column=14, value="はい" if same_face else "いいえ")
                
                current_row += 1
        
        # Auto-width columns
        for col in range(1, 15):
            ws.column_dimensions[chr(64 + col)].width = 11

    def add_position_details_sheet(self, trial_data: List[Dict]):
        """
        Add detailed position data sheet
        
        Args:
            trial_data: List of trial results
        """
        ws = self.wb.create_sheet("Position Details")
        
        # Headers
        headers = ['トライアル', 'エッジID', '初期X', '初期Y', '初期Z', 
                   '現在X', '現在Y', '現在Z', '移動距離', 'ペア間距離']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
        
        # Data rows
        current_row = 2
        for trial_idx, trial in enumerate(trial_data, 1):
            edges = trial['edges']
            pair_analyses = trial['pair_analyses']
            pair_distances = {p.pair_id: p.current_distance for p in pair_analyses}
            
            for edge in edges:
                ws.cell(row=current_row, column=1, value=trial_idx)
                ws.cell(row=current_row, column=2, value=edge.edge_id)
                ws.cell(row=current_row, column=3, value=round(edge.initial_position.x, 1))
                ws.cell(row=current_row, column=4, value=round(edge.initial_position.y, 1))
                ws.cell(row=current_row, column=5, value=round(edge.initial_position.z, 1))
                ws.cell(row=current_row, column=6, value=round(edge.current_position.x, 1))
                ws.cell(row=current_row, column=7, value=round(edge.current_position.y, 1))
                ws.cell(row=current_row, column=8, value=round(edge.current_position.z, 1))
                ws.cell(row=current_row, column=9, value=round(edge.moved_distance, 2))
                ws.cell(row=current_row, column=10, value=round(pair_distances.get(edge.pair_id, 0), 2))
                current_row += 1
        
        # Auto-width
        for col in range(1, 11):
            ws.column_dimensions[chr(64 + col)].width = 12
    
    def add_pattern_analysis_sheet(self, trial_data: List[Dict]):
        """
        Add pattern frequency analysis sheet
        
        Args:
            trial_data: List of trial results
        """
        ws = self.wb.create_sheet("Pattern Analysis")
        
        # Aggregate pattern data (統一基準に合わせる)
        distance_ranges = {
            '0-1.5 (near)': 0,
            '1.5-3.5 (medium)': 0,
            '3.5+ (far)': 0
        }
        
        same_face_count = 0
        different_face_count = 0
        
        for trial in trial_data:
            for pair in trial['pair_analyses']:
                # 統一基準を使用: near≤1.5, medium≤3.5, far>3.5
                if pair.distance_category == 'near':
                    distance_ranges['0-1.5 (near)'] += 1
                elif pair.distance_category == 'medium':
                    distance_ranges['1.5-3.5 (medium)'] += 1
                elif pair.distance_category == 'far':
                    distance_ranges['3.5+ (far)'] += 1
                else:
                    raise ValueError(f"Unknown distance category: {pair.distance_category}")
                
                if pair.same_face:
                    same_face_count += 1
                else:
                    different_face_count += 1
        
        # Headers
        ws['A1'] = "距離範囲分析"
        ws['A1'].font = Font(size=12, bold=True)
        
        row = 3
        ws['A3'] = "距離範囲"
        ws['B3'] = "頻度"
        ws['C3'] = "割合"
        ws['A3'].font = Font(bold=True)
        ws['B3'].font = Font(bold=True)
        ws['C3'].font = Font(bold=True)
        
        total = sum(distance_ranges.values())
        row = 4
        for range_name, count in distance_ranges.items():
            ws[f'A{row}'] = range_name
            ws[f'B{row}'] = count
            ws[f'C{row}'] = f"{count/total*100:.1f}%" if total > 0 else "0%"
            row += 1
        
        # Same face analysis
        row += 2
        ws[f'A{row}'] = "面の関係"
        ws[f'A{row}'].font = Font(size=12, bold=True)
        row += 2
        
        ws[f'A{row}'] = "同一面"
        ws[f'B{row}'] = same_face_count
        ws[f'C{row}'] = f"{same_face_count/(same_face_count+different_face_count)*100:.1f}%"
        row += 1
        
        ws[f'A{row}'] = "異なる面"
        ws[f'B{row}'] = different_face_count
        ws[f'C{row}'] = f"{different_face_count/(same_face_count+different_face_count)*100:.1f}%"
        
        # Column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
    
    def save(self, filename: str):
        """
        Save the workbook to file
        
        Args:
            filename: Output filename
        """
        self.wb.save(filename)
    
    def export_trial_data(self, trial_data: List[Dict], filename: str, 
                         operations_range: tuple = (15, 30)):
        """
        Export complete trial data to Excel
        
        Args:
            trial_data: List of trial results
            filename: Output filename
            operations_range: (min_ops, max_ops)
        """
        self.create_workbook()
        self.add_summary_sheet(trial_data, operations_range)
        self.add_pair_distances_sheet(trial_data)
        self.add_raw_data_sheet(trial_data)  # New: All edges for all trials
        self.add_position_details_sheet(trial_data)
        self.add_pattern_analysis_sheet(trial_data)
        self.save(filename)
        
        print(f"Excel file saved: {filename}")
        print(f"  - {len(trial_data)} trials exported")
        print(f"  - {len(trial_data) * 24} edge records in Raw Data sheet")
        print(f"  - 5 sheets created: Summary, Pair Distances, Raw Data, Position Details, Pattern Analysis")
    
    def export_csv(self, trial_data: List[Dict], filename: str):
        """
        Export raw data to CSV format for external analysis tools
        
        Args:
            trial_data: List of trial results
            filename: Output CSV filename
        """
        import csv
        
        with open(filename, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            
            # Headers
            writer.writerow(['Trial', 'Ops', 'Edge_ID', 'Pair_ID', 
                           'Init_X', 'Init_Y', 'Init_Z', 
                           'Curr_X', 'Curr_Y', 'Curr_Z', 
                           'Moved_Dist', 'Pair_Dist', 'Distance_Cat', 'Same_Face'])
            
            # Data rows
            for trial_idx, trial in enumerate(trial_data, 1):
                edges = trial['edges']
                pair_analyses = trial['pair_analyses']
                
                # Validate counts
                if len(edges) != 24:
                    raise ValueError(f"Trial {trial_idx} has {len(edges)} edges, expected 24")
                if len(pair_analyses) != 12:
                    raise ValueError(f"Trial {trial_idx} has {len(pair_analyses)} pairs, expected 12")
                
                # Create pair distance lookup
                pair_dist_dict = {p.pair_id: (p.current_distance, p.distance_category, p.same_face) 
                                 for p in pair_analyses}
                
                for edge in edges:
                    # Pair must exist - no default values
                    if edge.pair_id not in pair_dist_dict:
                        raise ValueError(f"Trial {trial_idx}: Pair {edge.pair_id} not found in analyses")
                    
                    pair_dist, dist_cat, same_face = pair_dist_dict[edge.pair_id]
                    
                    writer.writerow([
                        trial_idx,
                        trial['num_operations'],
                        edge.edge_id,
                        edge.pair_id,
                        round(edge.initial_position.x, 2),
                        round(edge.initial_position.y, 2),
                        round(edge.initial_position.z, 2),
                        round(edge.current_position.x, 2),
                        round(edge.current_position.y, 2),
                        round(edge.current_position.z, 2),
                        round(edge.moved_distance, 2),
                        round(pair_dist, 2),
                        dist_cat,
                        "Yes" if same_face else "No"
                    ])
        
        print(f"CSV file saved: {filename}")
        print(f"  - {len(trial_data)} trials exported")
        print(f"  - {len(trial_data) * 24} edge records")
    
    def export_statistical_analysis(self, stats_data: List[Dict], filename: str,
                                    min_ops: int, max_ops: int,
                                    trials_per_set: int, num_sets: int):
        """
        Export statistical analysis results to Excel
        
        Args:
            stats_data: List of statistical results per operation count
            filename: Output filename
            min_ops: Minimum operations
            max_ops: Maximum operations
            trials_per_set: Trials per set
            num_sets: Number of sets
        """
        self.create_workbook()
        
        # Create summary sheet
        ws = self.wb.create_sheet("Statistical Analysis", 0)
        
        # Title
        ws['A1'] = "4×4×4 Cube Analysis Simulator - 統計分析結果"
        ws['A1'].font = Font(size=14, bold=True)
        ws.merge_cells('A1:H1')
        
        # Parameters
        row = 3
        params = [
            ("操作数範囲:", f"{min_ops}-{max_ops}"),
            ("1セットあたりのトライアル数:", trials_per_set),
            ("セット反復回数:", num_sets),
            ("総トライアル数:", len(stats_data) * trials_per_set * num_sets),
            ("", ""),
        ]
        
        for label, value in params:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
        
        # Headers
        row += 1
        headers = [
            '操作数',
            '分離ペア数\n(平均)',
            '分離ペア数\n(標準偏差)',
            '平均ペア間距離\n(平均)',
            '平均ペア間距離\n(標準偏差)',
            '同一面ペア数\n(平均)',
            '同一面ペア数\n(標準偏差)'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Data rows
        for stat in stats_data:
            row += 1
            ws.cell(row=row, column=1, value=stat['num_operations'])
            ws.cell(row=row, column=2, value=round(stat['separated_pairs_mean'], 2))
            ws.cell(row=row, column=3, value=round(stat['separated_pairs_std'], 3))
            ws.cell(row=row, column=4, value=round(stat['avg_pair_distance_mean'], 2))
            ws.cell(row=row, column=5, value=round(stat['avg_pair_distance_std'], 3))
            ws.cell(row=row, column=6, value=round(stat['pairs_on_same_face_mean'], 2))
            ws.cell(row=row, column=7, value=round(stat['pairs_on_same_face_std'], 3))
            
            # Center align all data
            for col in range(1, 8):
                ws.cell(row=row, column=col).alignment = Alignment(horizontal='center', vertical='center')
        
        # Column widths
        ws.column_dimensions['A'].width = 10
        for col in ['B', 'C', 'D', 'E', 'F', 'G']:
            ws.column_dimensions[col].width = 15
        
        # Row height for header
        ws.row_dimensions[row - len(stats_data)].height = 40
        
        self.save(filename)
        
        print(f"Excel file saved: {filename}")
        print(f"  - {len(stats_data)} operation counts analyzed")
        print(f"  - {len(stats_data) * trials_per_set * num_sets} total trials")


